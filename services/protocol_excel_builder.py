"""
protocol_excel_builder.py

Parses the structured markdown output from protocol_service.py
and writes each section into its correct Excel sheet.

Sheet layout:
    Protocol_Info          — key-value fields + preconditions + environment
    Test_Case_Summary      — one row per test case
    Test_Steps             — one row per step
    Requirement_Traceability — requirement and risk traceability matrices
    Coverage_Summary       — metrics table
"""

import re
from io import BytesIO

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter


# ============================================================
# COLOUR PALETTE
# ============================================================

COLOUR = {
    "header_bg":      "1F4E79",   # dark navy
    "header_font":    "FFFFFF",   # white
    "subheader_bg":   "2E75B6",   # medium blue
    "subheader_font": "FFFFFF",
    "positive_bg":    "E2EFDA",   # light green
    "negative_bg":    "FCE4D6",   # light red/orange
    "boundary_bg":    "FFF2CC",   # light yellow
    "alarm_bg":       "FCE4D6",
    "workflow_bg":    "DDEBF7",   # light blue
    "error_bg":       "FCE4D6",
    "regression_bg":  "EAF0FB",
    "tc_row_bg":      "D6E4F0",   # test case row highlight
    "precond_bg":     "FFF9E6",
    "verify_bg":      "E8F5E9",   # very light green
    "cleanup_bg":     "F3E5F5",   # light purple
    "alt_row":        "F5F8FC",
}

SCENARIO_TYPE_COLOURS = {
    "Positive":       COLOUR["positive_bg"],
    "Negative":       COLOUR["negative_bg"],
    "Boundary":       COLOUR["boundary_bg"],
    "Alarm":          COLOUR["alarm_bg"],
    "Workflow":       COLOUR["workflow_bg"],
    "Error Handling": COLOUR["error_bg"],
    "Regression":     COLOUR["regression_bg"],
}

STEP_TYPE_COLOURS = {
    "Test Case":    COLOUR["tc_row_bg"],
    "Precondition": COLOUR["precond_bg"],
    "Verification": COLOUR["verify_bg"],
    "Cleanup":      COLOUR["cleanup_bg"],
}


# ============================================================
# STYLE HELPERS
# ============================================================

def _fill(hex_colour):
    return PatternFill(fill_type="solid", fgColor=hex_colour)


def _font(bold=False, colour="000000", size=10, italic=False):
    return Font(bold=bold, color=colour, size=size, italic=italic)


def _border():
    thin = Side(style="thin", color="BFBFBF")
    return Border(left=thin, right=thin, top=thin, bottom=thin)


def _align(wrap=True, horizontal="left", vertical="top"):
    return Alignment(wrap_text=wrap, horizontal=horizontal, vertical=vertical)


def _style_header_row(ws, row_num, num_cols):
    for col in range(1, num_cols + 1):
        cell = ws.cell(row=row_num, column=col)
        cell.fill = _fill(COLOUR["header_bg"])
        cell.font = _font(bold=True, colour=COLOUR["header_font"], size=10)
        cell.border = _border()
        cell.alignment = _align(horizontal="center", vertical="center")


def _style_data_cell(cell, bg=None, bold=False, italic=False):
    cell.font = _font(bold=bold, italic=italic)
    cell.border = _border()
    cell.alignment = _align()
    if bg:
        cell.fill = _fill(bg)


def _set_col_widths(ws, widths):
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _freeze(ws, cell_ref):
    ws.freeze_panes = cell_ref


# ============================================================
# MARKDOWN TABLE PARSER
# ============================================================

def _parse_markdown_table(text, section_marker):
    """
    Find a markdown table that appears after `section_marker` in `text`.
    Returns a list of lists (each inner list = one row's cell values).
    Skips the header row and separator row.
    """
    rows = []
    in_section = False
    found_header = False
    found_separator = False

    for line in text.splitlines():
        stripped = line.strip()

        # Detect section start
        if section_marker.lower() in stripped.lower():
            in_section = True
            continue

        if not in_section:
            continue

        # Stop at next major section (## heading)
        if re.match(r"^##\s+", stripped) and found_header:
            break

        if not stripped.startswith("|"):
            continue

        # Separator row
        if re.match(r"^\|[-| ]+\|$", stripped):
            found_separator = True
            continue

        # Header row (first pipe row before separator)
        if not found_separator:
            found_header = True
            continue

        # Data row
        cells = [c.strip() for c in stripped.strip("|").split("|")]
        rows.append(cells)

    return rows


def _parse_key_value_block(text, section_marker, stop_markers=None):
    """
    Extract key: value lines from a section.
    Returns a list of (key, value) tuples.
    """
    stop_markers = stop_markers or ["##", "---"]
    pairs = []
    in_section = False

    for line in text.splitlines():
        stripped = line.strip()

        if section_marker.lower() in stripped.lower() and stripped.startswith("#"):
            in_section = True
            continue

        if not in_section:
            continue

        # Stop conditions
        should_stop = any(stripped.startswith(s) for s in stop_markers)
        if should_stop and pairs:  # only stop if we already collected something
            break

        # key: value lines
        if ":" in stripped and not stripped.startswith("-") and not stripped.startswith("|"):
            parts = stripped.split(":", 1)
            key = parts[0].strip()
            val = parts[1].strip() if len(parts) > 1 else ""
            if key:
                pairs.append((key, val))

    return pairs


def _parse_bullet_list(text, section_marker, stop_markers=None):
    """
    Extract bullet list items from a section.
    Returns a list of strings.
    """
    stop_markers = stop_markers or ["##", "---"]
    items = []
    in_section = False

    for line in text.splitlines():
        stripped = line.strip()

        if section_marker.lower() in stripped.lower() and stripped.startswith("#"):
            in_section = True
            continue

        if not in_section:
            continue

        should_stop = any(stripped.startswith(s) for s in stop_markers)
        if should_stop and items:
            break

        if stripped.startswith("-") or stripped.startswith("•") or stripped.startswith("*"):
            item = stripped.lstrip("-•* ").strip()
            if item:
                items.append(item)

    return items


def _parse_coverage_metrics(text):
    """
    Extract coverage metric lines from COVERAGE SUMMARY section.
    Returns list of (metric, value) tuples.
    """
    metrics = []
    in_section = False
    skip_next = False

    for line in text.splitlines():
        stripped = line.strip()

        if "COVERAGE SUMMARY" in stripped.upper():
            in_section = True
            continue

        if not in_section:
            continue

        if stripped.startswith("##") and "COVERAGE" not in stripped.upper():
            break

        if not stripped or stripped.startswith("#") or stripped == "---":
            continue

        if ":" in stripped and not stripped.startswith("|"):
            parts = stripped.split(":", 1)
            key = parts[0].strip().lstrip("-• *")
            val = parts[1].strip()
            if key and val:
                metrics.append((key, val))

    return metrics


# ============================================================
# SHEET BUILDERS
# ============================================================

def _build_protocol_info_sheet(ws, protocol_text):
    ws.title = "Protocol_Info"

    # Title
    ws.merge_cells("A1:B1")
    title_cell = ws["A1"]
    title_cell.value = "ALTA Protocol — Protocol Information"
    title_cell.font = _font(bold=True, size=14, colour=COLOUR["header_font"])
    title_cell.fill = _fill(COLOUR["header_bg"])
    title_cell.alignment = _align(horizontal="center")
    ws.row_dimensions[1].height = 28

    # ── Key-value fields ──────────────────────────────────
    kv_pairs = _parse_key_value_block(protocol_text, "## PROTOCOL INFORMATION",
                                       stop_markers=["## PRECONDITIONS", "## TEST ENVIRONMENT", "---"])
    if not kv_pairs:
        # Try alternative heading
        kv_pairs = _parse_key_value_block(protocol_text, "PROTOCOL INFORMATION",
                                           stop_markers=["PRECONDITIONS", "TEST ENVIRONMENT", "---"])

    row = 3
    ws["A2"].value = "Field"
    ws["B2"].value = "Value"
    _style_header_row(ws, 2, 2)

    for key, val in kv_pairs:
        ws.cell(row=row, column=1).value = key
        ws.cell(row=row, column=2).value = val
        _style_data_cell(ws.cell(row=row, column=1), bold=True)
        _style_data_cell(ws.cell(row=row, column=2))
        row += 1

    row += 1  # blank row

    # ── Preconditions ─────────────────────────────────────
    ws.cell(row=row, column=1).value = "PRECONDITIONS"
    ws.cell(row=row, column=1).font = _font(bold=True, colour=COLOUR["header_font"])
    ws.cell(row=row, column=1).fill = _fill(COLOUR["subheader_bg"])
    ws.merge_cells(f"A{row}:B{row}")
    row += 1

    preconditions = _parse_bullet_list(protocol_text, "## PRECONDITIONS",
                                        stop_markers=["## TEST ENVIRONMENT", "---"])
    if not preconditions:
        preconditions = _parse_bullet_list(protocol_text, "PRECONDITIONS",
                                            stop_markers=["TEST ENVIRONMENT", "---"])

    for i, item in enumerate(preconditions):
        ws.cell(row=row, column=1).value = f"{i + 1}."
        ws.cell(row=row, column=2).value = item
        _style_data_cell(ws.cell(row=row, column=1))
        _style_data_cell(ws.cell(row=row, column=2))
        row += 1

    row += 1

    # ── Test Environment ──────────────────────────────────
    ws.cell(row=row, column=1).value = "TEST ENVIRONMENT"
    ws.cell(row=row, column=1).font = _font(bold=True, colour=COLOUR["header_font"])
    ws.cell(row=row, column=1).fill = _fill(COLOUR["subheader_bg"])
    ws.merge_cells(f"A{row}:B{row}")
    row += 1

    env_items = _parse_bullet_list(protocol_text, "## TEST ENVIRONMENT",
                                    stop_markers=["---", "## TEST CASE"])
    if not env_items:
        env_items = _parse_bullet_list(protocol_text, "TEST ENVIRONMENT",
                                        stop_markers=["---", "## TEST CASE"])

    for item in env_items:
        if ":" in item:
            parts = item.split(":", 1)
            ws.cell(row=row, column=1).value = parts[0].strip()
            ws.cell(row=row, column=2).value = parts[1].strip()
        else:
            ws.cell(row=row, column=1).value = item
            ws.cell(row=row, column=2).value = ""
        _style_data_cell(ws.cell(row=row, column=1), bold=True)
        _style_data_cell(ws.cell(row=row, column=2))
        row += 1

    _set_col_widths(ws, [30, 80])


def _build_test_case_summary_sheet(ws, protocol_text):
    ws.title = "Test_Case_Summary"

    headers = ["Execution Order", "Test Case ID", "Title", "Priority",
               "Scenario Type", "Related Requirements"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 20
    _freeze(ws, "A2")

    rows = _parse_markdown_table(protocol_text, "## TEST CASE SUMMARY")

    for i, row in enumerate(rows, start=2):
        # Pad to 6 columns
        while len(row) < 6:
            row.append("")

        for col_idx, val in enumerate(row[:6], start=1):
            cell = ws.cell(row=i, column=col_idx)
            cell.value = val
            # Colour by scenario type (column 5)
            scenario_type = row[4].strip() if len(row) > 4 else ""
            bg = SCENARIO_TYPE_COLOURS.get(scenario_type, None)
            _style_data_cell(cell, bg=bg)

        # Execution order — center-align
        ws.cell(row=i, column=1).alignment = _align(horizontal="center")
        ws.cell(row=i, column=4).alignment = _align(horizontal="center")

    _set_col_widths(ws, [14, 16, 55, 12, 16, 35])


def _build_test_steps_sheet(ws, protocol_text):
    ws.title = "Test_Steps"

    headers = ["Test Case ID", "Step ID", "Step Type", "Text",
               "Expected Result", "SRS ID", "Status", "Comment"]
    ws.append(headers)
    _style_header_row(ws, 1, len(headers))
    ws.row_dimensions[1].height = 20
    _freeze(ws, "A2")

    rows = _parse_markdown_table(protocol_text, "## TEST STEPS")

    for i, row in enumerate(rows, start=2):
        while len(row) < 8:
            row.append("")

        step_type = row[2].strip() if len(row) > 2 else ""
        bg = STEP_TYPE_COLOURS.get(step_type, None)

        # Bold the Test Case rows
        is_tc_row = step_type == "Test Case"

        for col_idx, val in enumerate(row[:8], start=1):
            cell = ws.cell(row=i, column=col_idx)
            cell.value = val
            _style_data_cell(cell, bg=bg, bold=is_tc_row)

        # Center step type and step ID columns
        ws.cell(row=i, column=1).alignment = _align(horizontal="center")
        ws.cell(row=i, column=2).alignment = _align(horizontal="center")
        ws.cell(row=i, column=3).alignment = _align(horizontal="center")
        ws.cell(row=i, column=6).alignment = _align(horizontal="center")

    _set_col_widths(ws, [14, 10, 14, 65, 55, 12, 10, 18])


def _build_traceability_sheet(ws, protocol_text):
    ws.title = "Requirement_Traceability"

    row_num = 1

    # ── Requirement Traceability ──────────────────────────
    ws.cell(row=row_num, column=1).value = "REQUIREMENT TRACEABILITY"
    ws.cell(row=row_num, column=1).font = _font(bold=True, colour=COLOUR["header_font"], size=11)
    ws.cell(row=row_num, column=1).fill = _fill(COLOUR["header_bg"])
    ws.merge_cells(f"A{row_num}:B{row_num}")
    row_num += 1

    ws.cell(row=row_num, column=1).value = "Requirement ID"
    ws.cell(row=row_num, column=2).value = "Mapped Test Cases"
    _style_header_row(ws, row_num, 2)
    row_num += 1

    req_rows = _parse_markdown_table(protocol_text, "## REQUIREMENT TRACEABILITY")
    for row in req_rows:
        while len(row) < 2:
            row.append("")
        ws.cell(row=row_num, column=1).value = row[0]
        ws.cell(row=row_num, column=2).value = row[1]
        _style_data_cell(ws.cell(row=row_num, column=1), bold=True)
        _style_data_cell(ws.cell(row=row_num, column=2))
        row_num += 1

    row_num += 1

    # ── Risk Traceability ─────────────────────────────────
    ws.cell(row=row_num, column=1).value = "RISK TRACEABILITY"
    ws.cell(row=row_num, column=1).font = _font(bold=True, colour=COLOUR["header_font"], size=11)
    ws.cell(row=row_num, column=1).fill = _fill(COLOUR["subheader_bg"])
    ws.merge_cells(f"A{row_num}:B{row_num}")
    row_num += 1

    ws.cell(row=row_num, column=1).value = "Risk ID"
    ws.cell(row=row_num, column=2).value = "Mapped Test Cases"
    _style_header_row(ws, row_num, 2)
    row_num += 1

    risk_rows = _parse_markdown_table(protocol_text, "## RISK TRACEABILITY")
    for row in risk_rows:
        while len(row) < 2:
            row.append("")
        ws.cell(row=row_num, column=1).value = row[0]
        ws.cell(row=row_num, column=2).value = row[1]
        _style_data_cell(ws.cell(row=row_num, column=1), bold=True)
        _style_data_cell(ws.cell(row=row_num, column=2))
        row_num += 1

    _set_col_widths(ws, [20, 60])
    _freeze(ws, "A2")


def _build_coverage_sheet(ws, protocol_text):
    ws.title = "Coverage_Summary"

    ws.cell(row=1, column=1).value = "COVERAGE SUMMARY"
    ws.cell(row=1, column=1).font = _font(bold=True, colour=COLOUR["header_font"], size=11)
    ws.cell(row=1, column=1).fill = _fill(COLOUR["header_bg"])
    ws.merge_cells("A1:B1")

    ws.cell(row=2, column=1).value = "Metric"
    ws.cell(row=2, column=2).value = "Value"
    _style_header_row(ws, 2, 2)

    metrics = _parse_coverage_metrics(protocol_text)

    for i, (metric, value) in enumerate(metrics, start=3):
        ws.cell(row=i, column=1).value = metric
        ws.cell(row=i, column=2).value = value
        _style_data_cell(ws.cell(row=i, column=1), bold=True)
        _style_data_cell(ws.cell(row=i, column=2))

    _set_col_widths(ws, [35, 30])


# ============================================================
# MAIN EXPORT FUNCTION
# ============================================================

def protocol_to_excel(protocol_text):
    """
    Convert the structured markdown protocol text into a formatted Excel workbook.
    Returns a BytesIO object ready to be downloaded.

    Args:
        protocol_text (str): Markdown output from protocol_service.generate_protocol()

    Returns:
        BytesIO: Excel workbook bytes
    """
    wb = Workbook()

    # Sheet 1 — Protocol Info
    ws_info = wb.active
    _build_protocol_info_sheet(ws_info, protocol_text)

    # Sheet 2 — Test Case Summary
    ws_summary = wb.create_sheet("Test_Case_Summary")
    _build_test_case_summary_sheet(ws_summary, protocol_text)

    # Sheet 3 — Test Steps
    ws_steps = wb.create_sheet("Test_Steps")
    _build_test_steps_sheet(ws_steps, protocol_text)

    # Sheet 4 — Traceability
    ws_trace = wb.create_sheet("Requirement_Traceability")
    _build_traceability_sheet(ws_trace, protocol_text)

    # Sheet 5 — Coverage Summary
    ws_cov = wb.create_sheet("Coverage_Summary")
    _build_coverage_sheet(ws_cov, protocol_text)

    output = BytesIO()
    wb.save(output)
    output.seek(0)
    return output